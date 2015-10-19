from __future__ import absolute_import
import daprot.mapper
from .. import (
    Reader as AbstractReader,
    Writer as AbstractWriter,
    exceptions
)
from io import BytesIO
from ..utils import *
from ..compat import *
from openpyxl import *


class Reader(AbstractReader):
    # void
    def __init__(self, resource, schema, offset=0, limit=None, sheet_name=None, merge_sheets=False, has_header=False, **fmtparams):
        self.fmtparams = fmtparams
        self.sheet_name = sheet_name
        # TODO merge functionality
        self.merge_sheets = merge_sheets
        self.has_header = has_header
        super(Reader, self).__init__(resource, schema, offset, limit)

    # function
    def get_mapper(self):
        return daprot.mapper.NAME

    def get_iterable_data(self):
        workbook = load_workbook(filename=BytesIO(encode(get_content(self.resource), self.encoding)))

        if self.merge_sheets:
            worksheet = workbook.worksheets
        else:
            worksheet = [workbook[self.sheet_name] if self.sheet_name else workbook.active]

        for ws in worksheet:
            rows = iter(ws.rows)
            if self.has_header:
                next(rows)
            for row in rows:
                yield [unicode(cell.value or u'') for cell in row]


class Writer(AbstractWriter):
    # void
    def __init__(self, filename, iterable_data, input_schema=None, add_header=True, sheet_name=None, replace_file = True, truncate_sheet = True, dinamic_sheet_field = None,  **fmtparams):
        self.fmtparams = fmtparams
        self.add_header = add_header
        self.sheet_name = sheet_name
        self.replace_file = replace_file
        self.truncate_sheet = truncate_sheet
        self.dinamic_sheet_field = dinamic_sheet_field
        self.is_exists_file = os.path.isfile(filename)
        self.is_exists_sheet = True
        super(Writer, self).__init__(filename, iterable_data, input_schema)

        if self.is_nested():
            raise exceptions.NestedSchemaNotSupported("{self} is not support nested schemas.".format(self=self.name))

    # openpyxl.Workbook
    def _get_workbook(self):
        if self.is_exists_file and not self.replace_file:
            return load_workbook(self.resource)
        return Workbook()

    # openpyxl.Worksheet
    def _get_worksheet(self, workbook):
        if not self.sheet_name:
            self.sheet_name = workbook.active.title
            return workbook.active

        if self.sheet_name in workbook.get_sheet_names():
            return workbook.get_sheet_by_name(self.sheet_name)

        self.is_exists_sheet = False
        worksheet = workbook.create_sheet(title=self.sheet_name)
        workbook.active = workbook.get_index(worksheet)
        return worksheet

    # tuple
    def init_writer(self):
        workbook = self._get_workbook()
        worksheet = self._get_worksheet(workbook)

        return workbook, worksheet

    # str in PY3 and unicode in PY2
    def unmarshal_item(self, item):
        if isinstance(item, (datetime.date, datetime.datetime)):
            return item.isoformat()

        # Convert to string.
        return unicode(item or u'')

    # void
    def write_header(self, worksheet):
        if self.add_header and worksheet.max_row == 1:
            worksheet.append(self.fieldnames)

    # void
    def _truncate_worksheet(self):
        if not self.truncate_sheet:
            return

        if not self.is_exists_file:
            return

        if not self.is_exists_sheet:
            return

        self.writer[0].remove_sheet(self.writer[1])
        self.writer = (self.writer[0], self.writer[0].create_sheet(title=self.sheet_name))

    # void
    def write(self):
        self._truncate_worksheet()
        self.write_header(self.writer[1])
        super(Writer, self).write()
        self.writer[0].save(self.resource)

    # openpyxl.Worksheet
    def _get_dinamic_sheet(self, worksheet_name):
        if worksheet_name in self.writer[0].get_sheet_names():
            return self.writer[0].get_sheet_by_name(worksheet_name)

        new_worksheet = self.writer[0].create_sheet(title=worksheet_name)
        self.write_header(new_worksheet)
        return new_worksheet

    # void
    def write_item(self, item):
        data = unmarshal(item, self.unmarshal_item)
        self.writer[1].append([data[c] for c in self.fieldnames])

        if not self.dinamic_sheet_field:
            return

        dinamic_sheet_name = data[self.dinamic_sheet_field]

        if dinamic_sheet_name is None:
            return

        dinamic_sheet = self._get_dinamic_sheet(dinamic_sheet_name)
        dinamic_sheet.append([data[c] for c in self.fieldnames])



