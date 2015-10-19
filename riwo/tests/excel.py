# -*- coding: utf-8 -*-

from __future__ import absolute_import
import os
import io
import shutil
from openpyxl import load_workbook
import riwo
import daprot as dp
import datetime
import unittest
import string
import random
from riwo.compat import *
from riwo.utils import *
from . import (
    CommonReader as AbstractCommonReader,
    CommonWriter as AbstractCommonWrite,
    __dir__, 
    __remote__
)


class CommonReader(AbstractCommonReader):
    expected_result = [{
      u'quantity': 1,
      u'name': u'đói',
      u'updated_at': u'2015.09.20 20:00',
      u'price': u'449',
      u'id': u'P0001'
    }, {
      u'quantity': 1,
      u'name': u'배고픈',
      u'updated_at': u'2015.09.20 20:02',
      u'price': u'399',
      u'id': u'P0002'
    }, {
      u'quantity': 10,
      u'name': u'голодный',
      u'updated_at': u'',
      u'price': u'199',
      u'id': u'P0003'
    }, {
      u'quantity': 1,
      u'name': u'Űrállomás krízis',
      u'updated_at': u'2015.09.20 12:47',
      u'price': u'999.5',
      u'id': u'P0004'
    }, {
      u'quantity': 1,
      u'name': u'Ovális iroda',
      u'updated_at': u'2015.09.20 07:31',
      u'price': u'2 399',
      u'id': u'P0005'
    }]


class CommonMergeReader(AbstractCommonReader):
    expected_result = [{
      u'quantity': 1,
      u'name': u'đói',
      u'updated_at': u'2015.09.20 20:00',
      u'price': u'449',
      u'id': u'P0001'
    }, {
      u'quantity': 1,
      u'name': u'배고픈',
      u'updated_at': u'2015.09.20 20:02',
      u'price': u'399',
      u'id': u'P0002'
    }, {
      u'quantity': 10,
      u'name': u'голодный',
      u'updated_at': u'',
      u'price': u'199',
      u'id': u'P0003'
    }, {
      u'quantity': 1,
      u'name': u'Űrállomás krízis',
      u'updated_at': u'2015.09.20 12:47',
      u'price': u'999.5',
      u'id': u'P0004'
    }, {
      u'quantity': 1,
      u'name': u'Ovális iroda',
      u'updated_at': u'2015.09.20 07:31',
      u'price': u'2 399',
      u'id': u'P0005'
    }]

class CommonWriter(AbstractCommonWrite):

    def setUp(self):
        self.test_file_path = self.test_file_base \
            .format(ext=self.ext, token=''.join([random.choice(string.ascii_uppercase) for i in range(6)]))
        self.resource = self.test_file_path

    def tearDown(self):
        if not self.create_file:
            self.assertFalse(os.path.isfile(self.test_file_path))
            return

        worksheet = load_workbook(self.test_file_path).active
        content = [[unicode(c.value or u'') for c in r] for r in worksheet.rows]
        self.assertEqual(self.content, content)
        os.remove(self.test_file_path)


class IndexSchema(dp.SchemaFlow):
    id = dp.Field(0)
    name = dp.Field(1, type=unicode, transforms=unicode.strip)
    price = dp.Field(2)
    quantity = dp.Field(3, type=long)
    updated_at = dp.Field(4)


class HeaderedSchema(dp.SchemaFlow):
    id = dp.Field()
    name = dp.Field(type=unicode, transforms=unicode.strip)
    price = dp.Field()
    quantity = dp.Field(type=long)
    updated_at = dp.Field()


class LocalReader(CommonReader, unittest.TestCase):
    def setUp(self):
        self.resource = io.open(os.path.join(__dir__, 'test.xlsx'), 'rb')
        self.reader = riwo.excel.Reader(self.resource, IndexSchema, sheet='Sheet1')


class LocalReaderMerge(CommonReader, unittest.TestCase):
    def setUp(self):
        self.resource = io.open(os.path.join(__dir__, 'test-merge.xlsx'), 'rb')
        self.reader = riwo.excel.Reader(self.resource, IndexSchema, merge_sheets=True)


class LocalReaderMergeWithHeader(CommonReader, unittest.TestCase):
    def setUp(self):
        self.resource = io.open(os.path.join(__dir__, 'test-merge-with-header.xlsx'), 'rb')
        self.reader = riwo.excel.Reader(self.resource, IndexSchema, merge_sheets=True, has_header=True)


# class RequestsReader(CommonReader, unittest.TestCase):
#     def setUp(self):
#         self.resource = requests.get(os.path.join(__remote__, 'test.xlsx'))
#         self.reader = riwo.excel.Reader(self.resource, IndexSchema, sheet='Sheet1')


class RequestsReaderMerge(CommonReader, unittest.TestCase):
    def setUp(self):
        self.resource = requests.get(os.path.join(__remote__, 'test-merge.xlsx'))
        self.reader = riwo.excel.Reader(self.resource, IndexSchema, merge_sheets=True)


class RequestsReaderMergeWithHeader(CommonReader, unittest.TestCase):
    def setUp(self):
        self.resource = requests.get(os.path.join(__remote__, 'test-merge-with-header.xlsx'))
        self.reader = riwo.excel.Reader(self.resource, IndexSchema, merge_sheets=True, has_header=True)


class UrllibReader(CommonReader, unittest.TestCase):
    def setUp(self):
        self.resource = urlopen(os.path.join(__remote__, 'test.xlsx'))
        self.reader = riwo.excel.Reader(self.resource, IndexSchema, sheet='Sheet1')


class UrllibReaderMerge(CommonReader, unittest.TestCase):
    def setUp(self):
        self.resource = urlopen(os.path.join(__remote__, 'test-merge.xlsx'))
        self.reader = riwo.excel.Reader(self.resource, IndexSchema, merge_sheets=True)


class UrllibReaderMergeWithHeader(CommonReader, unittest.TestCase):
    def setUp(self):
        self.resource = urlopen(os.path.join(__remote__, 'test-merge-with-header.xlsx'))
        self.reader = riwo.excel.Reader(self.resource, IndexSchema, merge_sheets=True, has_header=True)

class LocalWriter(CommonWriter, unittest.TestCase):
    class Schema(dp.SchemaFlow):
        id = dp.Field()
        name = dp.Field()
        price = dp.Field()
        quantity = dp.Field()
        updated_at = dp.Field()

    class NestedSchema(dp.SchemaFlow):
        inner_schema = dp.DictOf(IndexSchema)

    ext = 'xlsx'

    def test_add_header(self):
        self.writer = riwo.excel.Writer(self.resource, self.iterable_input, self.Schema, add_header=True)
        self.writer.write()
        self.create_file = True
        self.content = [
            [u'id', u'name', u'price', u'quantity', u'updated_at'],
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999,5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_not_add_header(self):
        self.writer = riwo.excel.Writer(self.resource, self.iterable_input, self.Schema, add_header=False)
        self.writer.write()
        self.create_file = True
        self.content = [
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999,5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_defined_schema_writer(self):
        self.writer = riwo.excel.Writer(self.resource, self.iterable_input, self.Schema)
        self.writer.write()
        self.create_file = True
        self.content = [
            [u'id', u'name', u'price', u'quantity', u'updated_at'],
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999,5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_schema_flow_writer(self):
        self.writer = riwo.excel.Writer(self.resource, self.Schema(self.iterable_input, mapper=dp.mapper.NAME))
        self.writer.write()
        self.create_file = True
        self.content = [
            [u'id', u'name', u'price', u'quantity', u'updated_at'],
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999,5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_without_schema(self):
        with self.assertRaises(riwo.exceptions.SchemaRequired):
            riwo.excel.Writer(self.resource, self.iterable_input)
        self.create_file = False
        self.content = u''

    def test_unmarshal(self):
        self.writer = riwo.excel.Writer(self.resource, [], self.Schema)
        self.create_file = False
        self.content = u''

        current_datetime = datetime.datetime.now()
        self.assertEqual(self.writer.unmarshal_item(current_datetime), current_datetime.isoformat())
        self.assertEqual(self.writer.unmarshal_item(4.21), u'4.21')
        self.assertEqual(self.writer.unmarshal_item(u'голодный'), u'голодный')

    def test_requisites(self):
        with self.assertRaises(riwo.exceptions.NestedSchemaNotSupported):
             riwo.excel.Writer(self.resource, [], self.NestedSchema)
        self.create_file = False
        self.content = u''


class CommonOptionWriter(AbstractCommonWrite):
    iterable_input = CommonReader.expected_result
    test_file_base = os.path.join(__dir__, 'output-{token}.{ext}')
    ext = 'xlsx'
    test_file_name = 'test.xlsx'
    sheet_name = None

    def setUp(self):
        self.test_file_path = self.test_file_base \
            .format(ext=self.ext, token=''.join([random.choice(string.ascii_uppercase) for i in range(6)]))
        shutil.copyfile(os.path.join(__dir__, self.test_file_name), self.test_file_path)
        self.resource = self.test_file_path
        self.dinamic_content = {}

    def tearDown(self):
        workbook = load_workbook(self.test_file_path)
        worksheet = workbook[self.sheet_name] if self.sheet_name else workbook.active
        content = [[unicode(c.value or u'') for c in r] for r in worksheet.rows]
        self.assertEqual(self.content, content)

        for sheet_name, sheet_content in list(self.dinamic_content.items()):
            self.assertEqual(
                sheet_content,
                [[unicode(c.value or u'') for c in r] for r in workbook[sheet_name].rows]
            )

        os.remove(self.test_file_path)


class LocalOptionWriter(CommonOptionWriter, unittest.TestCase):
    class Schema(dp.SchemaFlow):
        id = dp.Field()
        name = dp.Field()
        price = dp.Field()
        quantity = dp.Field()
        updated_at = dp.Field()

    test_file_name = 'test-option.xlsx'
    sheet_name = None

    def test_not_truncate_not_add_header(self):
        self.writer = riwo.excel.Writer(
            self.resource,
            self.iterable_input,
            self.Schema,
            sheet_name=self.sheet_name,
            add_header=False,
            truncate_sheet=False,
            replace_file=False
        )
        self.writer.write()
        self.content = [
            [u'P0006', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0007', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0008', u'голодный', u'199', u'10', u''],
            [u'P0009', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0010', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_not_truncate_add_header(self):
        self.writer = riwo.excel.Writer(
            self.resource,
            self.iterable_input,
            self.Schema,
            sheet_name=self.sheet_name,
            add_header=True,
            truncate_sheet=False,
            replace_file=False
        )
        self.writer.write()
        self.content = [
            [u'P0006', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0007', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0008', u'голодный', u'199', u'10', u''],
            [u'P0009', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0010', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_truncate_not_add_header(self):
        self.writer = riwo.excel.Writer(
            self.resource,
            self.iterable_input,
            self.Schema,
            sheet_name=self.sheet_name,
            add_header=False,
            truncate_sheet=True,
            replace_file=False
        )
        self.writer.write()
        self.content = [
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_truncate_add_header(self):
        self.writer = riwo.excel.Writer(
            self.resource,
            self.iterable_input,
            self.Schema,
            sheet_name=self.sheet_name,
            add_header=True,
            truncate_sheet=True,
            replace_file=False
        )
        self.writer.write()
        self.content = [
            [u'id', u'name', u'price', u'quantity', u'updated_at'],
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_replace_not_add_header(self):
        self.writer = riwo.excel.Writer(
            self.resource,
            self.iterable_input,
            self.Schema,
            sheet_name=self.sheet_name,
            add_header=False,
            truncate_sheet=False,
            replace_file=True
        )
        self.writer.write()
        self.content = [
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_replace_add_header(self):
        self.writer = riwo.excel.Writer(
            self.resource,
            self.iterable_input,
            self.Schema,
            sheet_name=self.sheet_name,
            add_header=True,
            truncate_sheet=False,
            replace_file=True
        )
        self.writer.write()
        self.content = [
            [u'id', u'name', u'price', u'quantity', u'updated_at'],
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]

    def test_dinamic_sheet_field_not_add_header(self):
        self.writer = riwo.excel.Writer(
            self.resource,
            self.iterable_input,
            self.Schema,
            sheet_name=self.sheet_name,
            add_header=False,
            truncate_sheet=False,
            replace_file=True,
            dinamic_sheet_field=u'quantity'
        )
        self.writer.write()
        self.content = [
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]
        self.dinamic_content = {
            u'1': [
                [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
                [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
                [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
                [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
            ],
            u'10': [
                [u'P0003', u'голодный', u'199', u'10', u''],
            ]
        }

    def test_dinamic_sheet_field_add_header(self):
        self.writer = riwo.excel.Writer(
            self.resource,
            self.iterable_input,
            self.Schema,
            sheet_name=self.sheet_name,
            add_header=True,
            truncate_sheet=False,
            replace_file=True,
            dinamic_sheet_field=u'quantity'
        )
        self.writer.write()
        self.content = [
            [u'id', u'name', u'price', u'quantity', u'updated_at'],
            [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
            [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
            [u'P0003', u'голодный', u'199', u'10', u''],
            [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
            [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
        ]
        self.dinamic_content = {
            u'1': [
                [u'id', u'name', u'price', u'quantity', u'updated_at'],
                [u'P0001', u'đói', u'449', u'1', u'2015.09.20 20:00'],
                [u'P0002', u'배고픈', u'399', u'1', u'2015.09.20 20:02'],
                [u'P0004', u'Űrállomás krízis', u'999.5', u'1', u'2015.09.20 12:47'],
                [u'P0005', u'Ovális iroda', u'2 399', u'1', u'2015.09.20 07:31'],
            ],
            u'10': [
                [u'id', u'name', u'price', u'quantity', u'updated_at'],
                [u'P0003', u'голодный', u'199', u'10', u''],
            ]
        }


class LocalOptionWriterWithSheetName(LocalOptionWriter, unittest.TestCase):

    sheet_name = u'Sheet1'